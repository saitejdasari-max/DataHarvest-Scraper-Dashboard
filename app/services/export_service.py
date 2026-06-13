import csv
import io
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.models.scraped_item import ScrapedItem


EXPORT_FIELDS = [
    "id", "title", "url", "source", "category",
    "company", "location", "salary", "job_type",
    "price", "currency", "rating", "review_count",
    "author", "published_at", "description", "scraped_at",
]


def items_to_rows(items: List[ScrapedItem]) -> List[dict]:
    rows = []
    for item in items:
        row = {}
        for field in EXPORT_FIELDS:
            val = getattr(item, field, None)
            if val is not None:
                row[field] = str(val) if not isinstance(val, (int, float)) else val
            else:
                row[field] = ""
        rows.append(row)
    return rows


def export_csv(items: List[ScrapedItem]) -> bytes:
    rows = items_to_rows(items)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=EXPORT_FIELDS)
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility


def export_excel(items: List[ScrapedItem]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scraped Data"

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Header row
    for col_idx, field in enumerate(EXPORT_FIELDS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    rows = items_to_rows(items)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, field in enumerate(EXPORT_FIELDS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row[field])
        # Zebra striping
        if row_idx % 2 == 0:
            for col_idx in range(1, len(EXPORT_FIELDS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color="F1F5F9", end_color="F1F5F9", fill_type="solid"
                )

    # Auto-fit columns
    for col_idx, field in enumerate(EXPORT_FIELDS, start=1):
        max_len = max(
            len(field),
            *(len(str(row[field])) for row in rows) if rows else [0],
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
